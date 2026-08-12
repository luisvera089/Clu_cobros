from __future__ import annotations

import ast
import hashlib
import json
import operator
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\Luis Vera\Downloads\Estado de pagos 2026 (1).xlsx")
OUTPUT = Path(__file__).resolve().parents[1] / "src" / "data" / "initial-data.json"
AS_OF_DATE = date.today()

TEAM_SHEETS = [
    "KIDS",
    "PIXIES",
    "MYSTIC",
    "MAGIC",
    "CRUSH",
    "WINGS",
    "CLAWS",
    "BLUE",
    "4EVER",
    "MA5",
    "FORMATIVO",
]

MONTH_INDEX = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

PREMIUM_TEAMS = {"BLUE", "4EVER", "MA5"}
COMPETITION_350_TEAMS = {"CLAWS", "CRUSH", "MAGIC", "PIXIES", "MYSTIC", "WINGS"}


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def normalize_key(value: str) -> str:
    text = clean_text(value).lower()
    text = "".join(
        char
        for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "sin-nombre"


def normalized_text(value: str) -> str:
    text = clean_text(value).lower()
    text = "".join(
        char
        for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def is_number(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def formula_value(formula: str):
    if not isinstance(formula, str) or not formula.startswith("="):
        return None
    expression = formula[1:].strip()
    if not re.fullmatch(r"[0-9+\-*/().\s]+", expression):
        return None

    operators = {
        ast.Add: operator.add,
        ast.Sub: operator.sub,
        ast.Mult: operator.mul,
        ast.Div: operator.truediv,
    }

    def evaluate(node):
        if isinstance(node, ast.Expression):
            return evaluate(node.body)
        if isinstance(node, ast.Num):
            return node.n
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return node.value
        if isinstance(node, ast.BinOp) and type(node.op) in operators:
            return operators[type(node.op)](evaluate(node.left), evaluate(node.right))
        if isinstance(node, ast.UnaryOp):
            if isinstance(node.op, ast.USub):
                return -evaluate(node.operand)
            if isinstance(node.op, ast.UAdd):
                return evaluate(node.operand)
        raise ValueError("unsupported formula")

    try:
        return float(evaluate(ast.parse(expression, mode="eval")))
    except Exception:
        return None


def cell_amount(ws_formula, ws_values, row: int, col: int):
    cached = ws_values.cell(row, col).value
    if is_number(cached):
        return float(cached)

    raw = ws_formula.cell(row, col).value
    if is_number(raw):
        return float(raw)

    evaluated = formula_value(raw)
    if evaluated is not None:
        return float(evaluated)

    return None


def header_info(ws, row: int):
    label = clean_text(ws.cell(row, 1).value).upper()
    if label == "DEPORTISTA":
        category = "Mensualidad" if row == 1 else "Concepto"
    elif "FLEXIBILIDAD" in label:
        category = "Flexibilidad"
    else:
        return None

    headers = []
    for col in range(2, ws.max_column + 1):
        header = clean_text(ws.cell(row, col).value)
        if not header:
            if headers:
                break
            continue
        if header.startswith("="):
            break
        headers.append({"col": col, "itemName": header})

    if not headers:
        return None
    return {"row": row, "category": category, "headers": headers}


def collect_sections(ws):
    sections = []
    row = 1
    while row <= ws.max_row:
        info = header_info(ws, row)
        if not info:
            row += 1
            continue

        end = row + 1
        while end <= ws.max_row and not header_info(ws, end):
            end += 1
        info["startRow"] = row + 1
        info["endRow"] = end - 1
        sections.append(info)
        row = end
    return sections


def mode(values):
    rounded = [int(round(v)) for v in values if is_number(v) and v > 0]
    if not rounded:
        return 0
    return Counter(rounded).most_common(1)[0][0]


def monthly_amount(month_no: int | None, paid_amount: float) -> int:
    base = 80000
    late = 90000
    if not month_no:
        return max(base, int(round(paid_amount or 0)))
    if month_no == AS_OF_DATE.month and AS_OF_DATE.day > 10 and paid_amount < base:
        return late
    if month_no > AS_OF_DATE.month:
        return base
    return max(base, int(round(paid_amount or 0)))


def expected_amount(team: str, athlete: str, category: str, item: str, paid_amount: float) -> tuple[int, dict]:
    item_norm = normalized_text(item)
    athlete_norm = normalized_text(athlete)
    paid = int(round(paid_amount or 0))
    meta = {}

    if category == "Mensualidad":
        month_no = MONTH_INDEX.get(clean_text(item).upper().strip())
        meta = {"monthNo": month_no, "baseAmount": 80000, "lateAmount": 90000, "lateDay": 10}
        return monthly_amount(month_no, paid_amount), meta

    if category == "Flexibilidad":
        return max(25000, paid), {"baseAmount": 25000}

    if "uniforme" in item_norm and "entrenamiento" in item_norm:
        if paid == 65000:
            return 65000, {"priceNote": "Una prenda"}
        return 130000, {"priceNote": "Conjunto completo"}

    if item_norm in {"mono", "mono"} or "mono" in item_norm or "moño" in clean_text(item).lower():
        return 30000, {}

    if "pista" in item_norm:
        if team in PREMIUM_TEAMS:
            return 50000, {}
        if team not in {"KIDS", "FORMATIVO"}:
            return 35000, {}
        return max(paid, 0), {}

    if "uni comp" in item_norm or ("uniforme" in item_norm and ("competencia" in item_norm or "comp" in item_norm)):
        if team in PREMIUM_TEAMS:
            return 400000, {}
        if team in COMPETITION_350_TEAMS:
            return 350000, {}
        return max(paid, 0), {}

    if "continental" in item_norm:
        return 190000, {}
    if "medcheer" in item_norm:
        return 100000, {}
    if "big show" in item_norm or "bigshow" in item_norm:
        return 100000, {}
    if item_norm == "liga" or "liga" in item_norm:
        return 100000, {}
    if "bucaramanga" in item_norm:
        return 720000, {}
    if "acompanante" in item_norm:
        quantity = paid if 0 < paid <= 10 else 1
        meta = {"unitAmount": 600000, "quantity": quantity}
        return int(600000 * quantity), meta
    if item_norm in {"capital", "capital bog"} or "capital bog" in item_norm:
        if "yeison betancur" in athlete_norm or athlete_norm == "yeison b":
            return 220000, {}
        return 420000, {}
    if "barranquilla" in item_norm:
        return 230000, {}
    if "highland" in item_norm or "higland" in item_norm or "bello" in item_norm:
        return 150000, {}
    if "finca" in item_norm:
        if team in {"MA5", "4EVER"}:
            return 100000, {}
        return max(paid, 0), {}

    return max(paid, 0), {}


def imported_payment_amount(category: str, item: str, amount: float | None) -> float:
    if not amount or amount <= 0:
        return 0
    item_norm = normalized_text(item)
    if "acompanante" in item_norm and amount <= 10:
        return 0
    return float(amount)


def build_initial_data():
    wb_formula = load_workbook(SOURCE, data_only=False)
    wb_values = load_workbook(SOURCE, data_only=True)

    teams = []
    default_samples = defaultdict(list)
    charge_records = []
    imported_submissions = []

    for team_name in TEAM_SHEETS:
        ws_formula = wb_formula[team_name]
        ws_values = wb_values[team_name]
        athlete_map = {}
        team = {"name": team_name, "id": normalize_key(team_name), "athletes": []}

        for section in collect_sections(ws_formula):
            category = section["category"]
            for row in range(section["startRow"], section["endRow"] + 1):
                athlete_name = clean_text(ws_formula.cell(row, 1).value)
                if not athlete_name:
                    continue
                if athlete_name.upper() in {"TOTAL", "TOTALES"}:
                    continue

                athlete_key = normalize_key(athlete_name)
                athlete_id = f"{team['id']}::{athlete_key}"
                if athlete_id not in athlete_map:
                    athlete_map[athlete_id] = {
                        "id": athlete_id,
                        "name": athlete_name,
                        "team": team_name,
                        "teamId": team["id"],
                        "charges": [],
                    }

                for header in section["headers"]:
                    item_name = clean_text(header["itemName"])
                    amount = cell_amount(ws_formula, ws_values, row, header["col"])
                    paid_amount = imported_payment_amount(category, item_name, amount)
                    total_amount, price_meta = expected_amount(team_name, athlete_name, category, item_name, paid_amount)
                    item_key = normalize_key(f"{category}-{item_name}")
                    charge_seed = f"{team_name}|{athlete_name}|{category}|{item_name}|{row}|{header['col']}"
                    charge_id = hashlib.sha1(charge_seed.encode("utf-8")).hexdigest()[:16]
                    source_cell = f"{team_name}!{ws_formula.cell(row, header['col']).coordinate}"

                    if paid_amount > 0:
                        default_samples[(team_name, category, item_name)].append(paid_amount)
                        default_samples[("*", category, item_name)].append(paid_amount)

                        payment_month = None
                        if category in {"Mensualidad", "Flexibilidad"}:
                            month_no = MONTH_INDEX.get(item_name.upper().strip())
                            if month_no:
                                payment_month = f"2026-{month_no:02d}-01"

                        imported_submissions.append(
                            {
                                "id": f"excel-{charge_id}",
                                "createdAt": "2026-07-15T00:00:00.000Z",
                                "paidAt": payment_month or "2026-07-15",
                                "athleteId": athlete_id,
                                "athleteName": athlete_name,
                                "team": team_name,
                                "payerName": "Importado del Excel",
                                "status": "importado",
                                "source": "Excel 2026",
                                "supportUrl": "",
                                "supportName": "",
                                "notes": source_cell,
                                "total": float(paid_amount),
                                "lines": [
                                    {
                                        "chargeId": charge_id,
                                        "category": category,
                                        "itemName": item_name,
                                        "amount": float(paid_amount),
                                        "detail": source_cell,
                                    }
                                ],
                            }
                        )

                    charge = {
                        "id": charge_id,
                        "category": category,
                        "itemName": item_name,
                        "itemKey": item_key,
                        "sourceCell": source_cell,
                        "excelAmount": float(paid_amount) if paid_amount > 0 else 0,
                        "rawExcelValue": float(amount) if amount and amount > 0 else 0,
                        "suggestedAmount": total_amount,
                        "expectedAmount": total_amount,
                        **price_meta,
                    }
                    athlete_map[athlete_id]["charges"].append(charge)
                    charge_records.append((team_name, category, item_name, charge))

        team["athletes"] = sorted(athlete_map.values(), key=lambda item: item["name"].lower())
        teams.append(team)

    for team_name, category, item_name, charge in charge_records:
        if not charge["suggestedAmount"]:
            charge["suggestedAmount"] = mode(default_samples[(team_name, category, item_name)])
        if not charge["suggestedAmount"]:
            charge["suggestedAmount"] = mode(default_samples[("*", category, item_name)])
        charge["expectedAmount"] = charge["suggestedAmount"]

    catalog = {}
    for team in teams:
        for athlete in team["athletes"]:
            for charge in athlete["charges"]:
                key = f"{charge['category']}::{charge['itemName']}"
                if key not in catalog:
                    catalog[key] = {
                        "category": charge["category"],
                        "itemName": charge["itemName"],
                        "suggestedAmount": charge["suggestedAmount"],
                    }

    for extra in [
        {"category": "Concepto", "itemName": "CAPITAL BARRANQUILLA", "suggestedAmount": 230000},
        {"category": "Concepto", "itemName": "HIGHLAND BELLO", "suggestedAmount": 150000},
        {"category": "Concepto", "itemName": "PRENDA UNIFORME", "suggestedAmount": 65000},
    ]:
        catalog.setdefault(f"{extra['category']}::{extra['itemName']}", extra)

    return {
        "generatedAt": date.today().isoformat(),
        "asOfDate": AS_OF_DATE.isoformat(),
        "sourceFile": str(SOURCE),
        "teams": teams,
        "catalog": sorted(catalog.values(), key=lambda item: (item["category"], item["itemName"])),
        "importedSubmissions": imported_submissions,
    }


def main():
    data = build_initial_data()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    athletes = sum(len(team["athletes"]) for team in data["teams"])
    charges = sum(len(athlete["charges"]) for team in data["teams"] for athlete in team["athletes"])
    print(f"Equipos: {len(data['teams'])}")
    print(f"Deportistas: {athletes}")
    print(f"Items asignados: {charges}")
    print(f"Pagos importados: {len(data['importedSubmissions'])}")
    print(f"Salida: {OUTPUT}")


if __name__ == "__main__":
    main()

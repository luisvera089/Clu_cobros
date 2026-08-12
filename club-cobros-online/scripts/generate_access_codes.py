from __future__ import annotations

import csv
import json
import secrets
import unicodedata
from datetime import datetime
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
INITIAL_DATA = ROOT / "src" / "data" / "initial-data.json"
STORE = ROOT / "data" / "store.json"
ACCESS_CODES = ROOT / "data" / "access-codes.json"
OUTPUT_CSV = ROOT.parents[1] / "outputs" / "codigos-acceso-deportistas.csv"
OUTPUT_XLSX = ROOT.parents[1] / "outputs" / "codigos-acceso-deportistas.xlsx"


def team_code(value: str) -> str:
    text = "".join(
        char
        for char in unicodedata.normalize("NFKD", value or "EQ")
        if not unicodedata.combining(char)
    )
    text = "".join(char for char in text.upper() if char.isalnum())
    return (text[:5] or "EQ")


def new_code(team: str, serial: int) -> str:
    return f"MAS-{team_code(team)}-{serial:03d}-{secrets.token_hex(3).upper()}"


def load_json(path: Path, fallback):
    if not path.exists():
        return fallback
    return json.loads(path.read_text(encoding="utf-8"))


def active_athletes():
    initial = load_json(INITIAL_DATA, {})
    store = load_json(STORE, {})
    removed = set(store.get("removedAthleteIds") or [])
    athletes = []
    for team in initial.get("teams", []):
        for athlete in team.get("athletes", []):
            if athlete["id"] not in removed:
                athletes.append(
                    {
                        "athleteId": athlete["id"],
                        "athleteName": athlete["name"],
                        "team": athlete["team"],
                    }
                )
    for athlete in store.get("customAthletes") or []:
        if athlete.get("active", True) and athlete["id"] not in removed:
            athletes.append(
                {
                    "athleteId": athlete["id"],
                    "athleteName": athlete["name"],
                    "team": athlete["team"],
                }
            )
    return sorted(athletes, key=lambda item: (item["team"], item["athleteName"].lower()))


def write_excel(rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        print("openpyxl no esta disponible; se omitio el archivo XLSX.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Codigos"

    headers = ["Equipo", "Deportista", "Codigo inicial", "Debe cambiar contrasena", "ID deportista"]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["team"],
                row["athleteName"],
                row["accessCode"],
                "SI" if row.get("mustChangePassword", True) else "NO",
                row["athleteId"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="0F2A44")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=thin)

    for column, width in {"A": 16, "B": 32, "C": 24, "D": 26, "E": 34}.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if ws.max_row >= 2:
        table = Table(displayName="TablaCodigosAcceso", ref=f"A1:E{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    info = wb.create_sheet("Instrucciones")
    info.append(["Archivo de codigos de acceso"])
    info.append(["Generado", datetime.now().strftime("%Y-%m-%d %H:%M")])
    info.append(["Uso", "Entrega a cada familia solo el codigo de su deportista."])
    info.append(["Primer ingreso", "La aplicacion obliga a cambiar el codigo inicial por una nueva contrasena."])
    info.append(["Administrador", "Desde el panel administrador puedes cambiar o reiniciar codigos."])
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 80
    info["A1"].fill = header_fill
    info["A1"].font = Font(color="FFFFFF", bold=True, size=14)

    wb.save(OUTPUT_XLSX)


def main():
    ACCESS_CODES.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)

    current = load_json(ACCESS_CODES, [])
    by_id = {item["athleteId"]: item for item in current}
    athletes = active_athletes()
    active_ids = {item["athleteId"] for item in athletes}
    serial = len(current) + 1

    for athlete in athletes:
      existing = by_id.get(athlete["athleteId"])
      if existing:
          existing["athleteName"] = athlete["athleteName"]
          existing["team"] = athlete["team"]
          existing["active"] = True
          if "mustChangePassword" not in existing:
              existing["mustChangePassword"] = True
      else:
          record = {
              **athlete,
              "accessCode": new_code(athlete["team"], serial),
              "mustChangePassword": True,
              "active": True,
          }
          serial += 1
          current.append(record)
          by_id[record["athleteId"]] = record

    for record in current:
        if record["athleteId"] not in active_ids:
            record["active"] = False

    ACCESS_CODES.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")

    rows = [item for item in current if item.get("active", True)]
    rows.sort(key=lambda item: (item["team"], item["athleteName"].lower()))
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["team", "athleteName", "accessCode", "mustChangePassword", "athleteId"])
        writer.writeheader()
        writer.writerows(
            {
                "team": row["team"],
                "athleteName": row["athleteName"],
                "accessCode": row["accessCode"],
                "mustChangePassword": row.get("mustChangePassword", True),
                "athleteId": row["athleteId"],
            }
            for row in rows
        )
    write_excel(rows)

    print(f"Codigos activos: {len(rows)}")
    print(f"Servidor: {ACCESS_CODES}")
    print(f"Archivo CSV para asignar: {OUTPUT_CSV}")
    print(f"Archivo Excel para asignar: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
